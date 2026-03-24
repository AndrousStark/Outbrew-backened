"""
Intelligent CSV/XLSX Parser Service

Ultra-intelligent parser with:
- Fuzzy column matching (handles typos, variations, different languages)
- Multi-sheet XLSX support
- Intelligent name merging (first + last → full name)
- Country detection and localization
- Data validation and confidence scoring
- Comprehensive error handling
"""

import csv
import io
import re
import logging
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass, field
from fuzzywuzzy import fuzz, process
import openpyxl
import chardet
import pandas as pd

logger = logging.getLogger(__name__)


@dataclass
class ColumnMapping:
    """Column mapping with confidence score"""
    detected_column: str
    mapped_to: str
    confidence: float  # 0-100
    original_values_sample: List[str] = field(default_factory=list)


@dataclass
class ParsedRecipient:
    """Parsed recipient with validation"""
    email: Optional[str] = None
    name: Optional[str] = None
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    company: Optional[str] = None
    position: Optional[str] = None
    country: Optional[str] = None
    city: Optional[str] = None
    linkedin_url: Optional[str] = None
    website: Optional[str] = None
    phone: Optional[str] = None

    # Metadata
    row_number: int = 0
    is_valid: bool = True
    validation_errors: List[str] = field(default_factory=list)
    confidence_score: float = 0.0  # 0-100

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for API response"""
        return {
            "email": self.email,
            "name": self.name,
            "first_name": self.first_name,
            "last_name": self.last_name,
            "company": self.company,
            "position": self.position,
            "country": self.country,
            "city": self.city,
            "linkedin_url": self.linkedin_url,
            "website": self.website,
            "phone": self.phone,
            "row_number": self.row_number,
            "is_valid": self.is_valid,
            "validation_errors": self.validation_errors,
            "confidence_score": self.confidence_score
        }


@dataclass
class ParseResult:
    """Complete parse result with statistics"""
    recipients: List[ParsedRecipient]
    column_mappings: List[ColumnMapping]
    detected_country: Optional[str] = None
    country_confidence: float = 0.0
    total_rows: int = 0
    valid_rows: int = 0
    invalid_rows: int = 0
    warnings: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for API response"""
        return {
            "recipients": [r.to_dict() for r in self.recipients],
            "column_mappings": [
                {
                    "detected_column": m.detected_column,
                    "mapped_to": m.mapped_to,
                    "confidence": m.confidence,
                    "original_values_sample": m.original_values_sample[:3]  # First 3 samples
                }
                for m in self.column_mappings
            ],
            "detected_country": self.detected_country,
            "country_confidence": self.country_confidence,
            "total_rows": self.total_rows,
            "valid_rows": self.valid_rows,
            "invalid_rows": self.invalid_rows,
            "warnings": self.warnings
        }


class IntelligentCSVParser:
    """
    Ultra-intelligent CSV/XLSX parser with fuzzy matching and validation.

    Features:
    - Handles 100+ column name variations
    - Multi-sheet XLSX support
    - Intelligent name merging
    - Country detection
    - Data validation
    - Confidence scoring
    """

    # Column patterns with fuzzy matching (lowercase for matching)
    EMAIL_PATTERNS = [
        "email", "mail", "e-mail", "e mail", "email address", "mail address",
        "email id", "mail id", "e-mail address", "electronic mail", "emails",
        "e_mail", "email_address", "mail_address", "correo", "courriel", "epost"
    ]

    FIRST_NAME_PATTERNS = [
        "first name", "firstname", "first_name", "fname", "given name",
        "givenname", "given_name", "forename", "prenom", "vorname", "first"
    ]

    LAST_NAME_PATTERNS = [
        "last name", "lastname", "last_name", "lname", "surname", "family name",
        "familyname", "family_name", "nom", "nachname", "last", "sur name"
    ]

    FULL_NAME_PATTERNS = [
        "name", "full name", "fullname", "full_name", "recipient name",
        "contact name", "person name", "nom complet", "vollstandiger name",
        "recipient", "contact", "person", "full name"
    ]

    COMPANY_PATTERNS = [
        "company", "comapny", "compnay", "organization", "organisation",
        "org", "company name", "companyname", "company_name",
        "organization name", "organizationname", "organization_name",
        "organisation name", "organisationname", "employer", "business",
        "enterprise", "firma", "entreprise", "societe", "unternehmen",
        "organizationname", "company/organization", "org name", "orgname"
    ]

    # Organization website patterns (should NOT match to company)
    ORGANIZATION_WEBSITE_PATTERNS = [
        "organizationwebsite", "organization website", "company website",
        "org website", "organizationlinkedinurl", "organization linkedin url",
        "org linkedin", "company linkedin", "organization linkedin", "org url"
    ]

    POSITION_PATTERNS = [
        "position", "title", "job title", "jobtitle", "job_title",
        "role", "job role", "designation", "job", "occupation",
        "fonction", "poste", "position", "stelle", "job position"
    ]

    COUNTRY_PATTERNS = [
        "country", "nation", "state", "country name", "land", "pays", "pais"
    ]

    CITY_PATTERNS = [
        "city", "town", "location", "place", "ville", "stadt", "ciudad"
    ]

    LINKEDIN_PATTERNS = [
        "linkedin", "linkedin url", "linkedinurl", "linkedin_url",
        "linkedin profile", "linkedin link", "profile url", "profile",
        "linkedin address", "li url", "li_url",
        "organizationlinkedinurl", "organization linkedin url", "org linkedin url",
        "company linkedin url", "company linkedin"
    ]

    WEBSITE_PATTERNS = [
        "website", "web site", "web", "site", "url", "homepage",
        "company website", "company site", "organization website",
        "organizationwebsite", "organization_website", "web address",
        "site web", "webseite", "org website", "orgwebsite"
    ]

    PHONE_PATTERNS = [
        "phone", "telephone", "phone number", "tel", "mobile",
        "cell", "contact number", "telephone number", "telefon",
        "telefono", "phone_number", "tel_number"
    ]

    # Country-specific email best practices
    COUNTRY_EMAIL_GUIDANCE = {
        "Luxembourg": {
            "formality": "formal",
            "language": "en",  # English widely used in business
            "tips": [
                "Use professional titles and surnames",
                "Keep emails concise and structured",
                "Multilingual context (French, German, English) - English is safe",
                "Emphasize international experience"
            ]
        },
        "Germany": {
            "formality": "very_formal",
            "language": "de",
            "tips": [
                "Always use 'Sehr geehrte/r' (Dear) + title + last name",
                "Include urgency in subject line ('Action Needed')",
                "Be punctual and detail-oriented in communication",
                "Avoid casual language like 'Hi' or 'Hey'",
                "Structure: clear, direct, comprehensive"
            ]
        },
        "Switzerland": {
            "formality": "very_formal",
            "language": "de",  # Depends on region (DE/FR/IT)
            "tips": [
                "Similar to Germany - very formal approach",
                "Include urgency levels in subject",
                "Punctuality is critical",
                "Address by title and last name",
                "Consider language region (German/French/Italian)"
            ]
        },
        "Poland": {
            "formality": "formal",
            "language": "pl",
            "tips": [
                "Use formal titles (Pan/Pani + surname)",
                "Building relationships is important",
                "Show respect for hierarchy",
                "GDPR: Requires prior consent for cold emails"
            ]
        },
        "Denmark": {
            "formality": "semi_formal",
            "language": "da",
            "tips": [
                "More relaxed than Germany but still professional",
                "First names acceptable after initial contact",
                "Direct and transparent communication style",
                "GDPR: Check Robinson List before contacting"
            ]
        },
        "Ireland": {
            "formality": "friendly_professional",
            "language": "en",
            "tips": [
                "Warmer, more conversational tone acceptable",
                "Relationship-building emphasized",
                "Humor can be appropriate when done tastefully",
                "GDPR: Requires opt-in consent for cold emails"
            ]
        },
        "Singapore": {
            "formality": "formal",
            "language": "en",
            "tips": [
                "Respect hierarchical structures",
                "Be concise and efficient",
                "Avoid aggressive sales tactics",
                "English is primary business language"
            ]
        },
        "USA": {
            "formality": "semi_formal",
            "language": "en",
            "tips": [
                "Direct and action-oriented communication",
                "Value proposition should be clear upfront",
                "Personalization is highly valued",
                "Follow-ups expected and acceptable"
            ]
        },
        "Canada": {
            "formality": "friendly_professional",
            "language": "en",  # en or fr depending on region
            "tips": [
                "Polite and inclusive communication",
                "Consider bilingual context (English/French)",
                "Similar to USA but slightly more formal",
                "Emphasize diversity and inclusion"
            ]
        },
        "Australia": {
            "formality": "casual_professional",
            "language": "en",
            "tips": [
                "More casual and friendly tone acceptable",
                "Avoid excessive formality",
                "Direct communication appreciated",
                "Use first names readily"
            ]
        }
    }

    def __init__(self, fuzzy_threshold: int = 70):
        """
        Initialize parser.

        Args:
            fuzzy_threshold: Minimum fuzzy match score (0-100) to consider a match
        """
        self.fuzzy_threshold = fuzzy_threshold

    def parse_file(
        self,
        file_content: bytes,
        filename: str,
        sheet_name: Optional[str] = None
    ) -> ParseResult:
        """
        Parse CSV or XLSX file with intelligent column detection.

        Args:
            file_content: Raw file bytes
            filename: Original filename
            sheet_name: Specific sheet to parse (XLSX only)

        Returns:
            ParseResult with parsed data and metadata
        """
        logger.info(f"🔍 [INTELLIGENT PARSER] Starting parse of: {filename}")

        # Detect file type
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            return self._parse_excel(file_content, sheet_name)
        else:
            # Try to detect encoding
            detected_encoding = chardet.detect(file_content)
            encoding = detected_encoding.get('encoding', 'utf-8')
            logger.info(f"📝 [PARSER] Detected encoding: {encoding}")

            # Decode and parse CSV
            try:
                text_content = file_content.decode(encoding)
            except Exception:
                text_content = file_content.decode('utf-8', errors='replace')

            return self._parse_csv(text_content)

    def _parse_excel(
        self,
        file_content: bytes,
        sheet_name: Optional[str] = None
    ) -> ParseResult:
        """Parse Excel file (XLSX/XLS)"""
        logger.info("📊 [PARSER] Parsing Excel file...")

        # Load workbook
        workbook = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)

        # Get sheet
        if sheet_name and sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            logger.info(f"📄 [PARSER] Using sheet: {sheet_name}")
        else:
            sheet = workbook.active
            logger.info(f"📄 [PARSER] Using active sheet: {sheet.title}")

        # Convert to CSV format for unified processing
        csv_data = []
        for row in sheet.iter_rows(values_only=True):
            csv_data.append([str(cell) if cell is not None else "" for cell in row])

        # Create CSV string
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerows(csv_data)
        csv_content = output.getvalue()

        return self._parse_csv(csv_content)

    def _parse_csv(self, csv_content: str) -> ParseResult:
        """Parse CSV content with intelligent column mapping"""
        logger.info("🧠 [PARSER] Starting intelligent CSV parsing...")

        # Parse CSV
        csv_reader = csv.DictReader(io.StringIO(csv_content))
        headers = csv_reader.fieldnames

        if not headers:
            return ParseResult(
                recipients=[],
                column_mappings=[],
                warnings=["No headers found in file"]
            )

        logger.info(f"📋 [PARSER] Found {len(headers)} columns: {headers}")

        # Map columns intelligently
        column_mappings = self._map_columns(headers, csv_reader)

        # Reset reader
        csv_reader = csv.DictReader(io.StringIO(csv_content))

        # Parse recipients
        recipients = []
        row_num = 1  # Start at 1 (header is 0)

        for row in csv_reader:
            row_num += 1
            recipient = self._parse_row(row, column_mappings, row_num)
            recipients.append(recipient)

        # Detect country from data
        detected_country, country_confidence = self._detect_country(recipients)

        # Calculate statistics
        valid_recipients = [r for r in recipients if r.is_valid]
        invalid_recipients = [r for r in recipients if not r.is_valid]

        # Generate warnings
        warnings = []
        if invalid_recipients:
            warnings.append(f"{len(invalid_recipients)} rows have validation errors")

        missing_emails = [r for r in recipients if not r.email]
        if missing_emails:
            warnings.append(f"{len(missing_emails)} rows are missing email addresses")

        result = ParseResult(
            recipients=recipients,
            column_mappings=column_mappings,
            detected_country=detected_country,
            country_confidence=country_confidence,
            total_rows=len(recipients),
            valid_rows=len(valid_recipients),
            invalid_rows=len(invalid_recipients),
            warnings=warnings
        )

        logger.info(
            f"✅ [PARSER] Parsing complete: {result.valid_rows}/{result.total_rows} valid rows"
        )

        return result

    def _map_columns(
        self,
        headers: List[str],
        csv_reader: csv.DictReader
    ) -> List[ColumnMapping]:
        """
        Intelligently map CSV columns to our schema using fuzzy matching.
        """
        logger.info("🎯 [PARSER] Mapping columns with fuzzy matching...")

        mappings = []

        # Read first few rows for sample values
        sample_rows = []
        for i, row in enumerate(csv_reader):
            if i >= 5:  # Get first 5 rows as sample
                break
            sample_rows.append(row)

        # Define all pattern groups
        pattern_groups = {
            "email": self.EMAIL_PATTERNS,
            "first_name": self.FIRST_NAME_PATTERNS,
            "last_name": self.LAST_NAME_PATTERNS,
            "full_name": self.FULL_NAME_PATTERNS,
            "company": self.COMPANY_PATTERNS,
            "position": self.POSITION_PATTERNS,
            "country": self.COUNTRY_PATTERNS,
            "city": self.CITY_PATTERNS,
            "linkedin_url": self.LINKEDIN_PATTERNS,
            "website": self.WEBSITE_PATTERNS,
            "phone": self.PHONE_PATTERNS
        }

        # Map each header
        for header in headers:
            best_match = None
            best_score = 0
            best_field = None

            header_lower = header.lower().strip()

            # Try fuzzy matching against all pattern groups
            for field_name, patterns in pattern_groups.items():
                match = process.extractOne(
                    header_lower,
                    patterns,
                    scorer=fuzz.ratio
                )

                if match and match[1] > best_score:
                    best_score = match[1]
                    best_match = match[0]
                    best_field = field_name

            # Only map if confidence is above threshold
            if best_score >= self.fuzzy_threshold:
                # Get sample values
                sample_values = [
                    row.get(header, "")[:50]  # First 50 chars
                    for row in sample_rows
                    if row.get(header)
                ]

                mapping = ColumnMapping(
                    detected_column=header,
                    mapped_to=best_field,
                    confidence=best_score,
                    original_values_sample=sample_values
                )
                mappings.append(mapping)

                logger.info(
                    f"  ✓ '{header}' → '{best_field}' (confidence: {best_score}%)"
                )
            else:
                logger.warning(
                    f"  ⚠ '{header}' - no confident match (best: {best_score}%)"
                )

        return mappings

    def _parse_row(
        self,
        row: Dict[str, str],
        mappings: List[ColumnMapping],
        row_number: int
    ) -> ParsedRecipient:
        """Parse a single row using column mappings"""
        recipient = ParsedRecipient(row_number=row_number)

        # Create reverse mapping (original column → field)
        mapping_dict = {m.detected_column: m.mapped_to for m in mappings}

        # Extract values
        for column, value in row.items():
            field = mapping_dict.get(column)
            if field and value:
                value = value.strip()
                setattr(recipient, field, value)

        # Intelligent name merging
        if not recipient.full_name and recipient.first_name and recipient.last_name:
            recipient.name = f"{recipient.first_name} {recipient.last_name}"
        elif not recipient.full_name and recipient.first_name:
            recipient.name = recipient.first_name
        elif recipient.full_name:
            recipient.name = recipient.full_name

        # Split full name if we have it but not first/last
        if recipient.name and not recipient.first_name:
            parts = recipient.name.split()
            if len(parts) >= 2:
                recipient.first_name = parts[0]
                recipient.last_name = " ".join(parts[1:])
            elif len(parts) == 1:
                recipient.first_name = parts[0]

        # Validate
        recipient.is_valid, recipient.validation_errors, recipient.confidence_score = (
            self._validate_recipient(recipient)
        )

        return recipient

    def _validate_recipient(
        self,
        recipient: ParsedRecipient
    ) -> Tuple[bool, List[str], float]:
        """
        Validate recipient and calculate confidence score.

        Returns:
            Tuple of (is_valid, errors, confidence_score)
        """
        errors = []
        confidence = 100.0

        # Email validation
        if not recipient.email:
            errors.append("Missing email address")
            confidence -= 50
        elif not self._is_valid_email(recipient.email):
            errors.append(f"Invalid email format: {recipient.email}")
            confidence -= 30

        # Name validation
        if not recipient.name and not (recipient.first_name or recipient.last_name):
            errors.append("Missing name")
            confidence -= 20

        # Company validation
        if not recipient.company:
            errors.append("Missing company name")
            confidence -= 10

        # LinkedIn URL validation
        if recipient.linkedin_url and not self._is_valid_linkedin_url(recipient.linkedin_url):
            errors.append("Invalid LinkedIn URL format")
            confidence -= 5

        # Website URL validation
        if recipient.website and not self._is_valid_url(recipient.website):
            errors.append("Invalid website URL format")
            confidence -= 5

        is_valid = len(errors) == 0 or (recipient.email and self._is_valid_email(recipient.email))
        confidence = max(0, confidence)

        return is_valid, errors, confidence

    def _is_valid_email(self, email: str) -> bool:
        """Validate email format"""
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return bool(re.match(pattern, email))

    def _is_valid_linkedin_url(self, url: str) -> bool:
        """Validate LinkedIn URL"""
        return 'linkedin.com' in url.lower()

    def _is_valid_url(self, url: str) -> bool:
        """Validate URL format"""
        pattern = r'^https?://[^\s<>"]+|www\.[^\s<>"]+'
        return bool(re.match(pattern, url))

    def _detect_country(
        self,
        recipients: List[ParsedRecipient]
    ) -> Tuple[Optional[str], float]:
        """
        Detect primary country from recipients data.

        Returns:
            Tuple of (country_name, confidence_score)
        """
        # Count country occurrences
        country_counts = {}
        total_with_country = 0

        for recipient in recipients:
            if recipient.country:
                country = recipient.country.strip().title()
                country_counts[country] = country_counts.get(country, 0) + 1
                total_with_country += 1

        if not country_counts:
            return None, 0.0

        # Get most common country
        most_common_country = max(country_counts, key=country_counts.get)
        count = country_counts[most_common_country]

        # Calculate confidence
        confidence = (count / len(recipients)) * 100

        logger.info(
            f"🌍 [PARSER] Detected country: {most_common_country} "
            f"({count}/{len(recipients)} = {confidence:.1f}%)"
        )

        return most_common_country, confidence

    def get_country_guidance(self, country: str) -> Optional[Dict[str, Any]]:
        """
        Get email writing guidance for specific country.

        Args:
            country: Country name

        Returns:
            Dictionary with formality level, language, and tips
        """
        return self.COUNTRY_EMAIL_GUIDANCE.get(country)
