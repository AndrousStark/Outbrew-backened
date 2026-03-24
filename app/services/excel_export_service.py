"""
Excel Export Service
FREE - Generate formatted Excel files from extraction results
Uses openpyxl for professional styling
"""

import os
from typing import List, Dict, Any, Optional
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.extraction import ExtractionJob, ExtractionResult
from app.core.config import settings

import logging

logger = logging.getLogger(__name__)


class ExcelExportService:
    """
    Export extraction results to formatted Excel files
    FREE - no API costs
    """

    def __init__(self, db: Session):
        self.db = db

    def export_job_results(
        self,
        job_id: int,
        include_metadata: bool = True,
        min_quality: Optional[float] = None
    ) -> str:
        """
        Export extraction results to Excel file

        Args:
            job_id: Extraction job ID
            include_metadata: Include quality scores and metadata columns
            min_quality: Minimum quality score filter (0-1)

        Returns:
            File path to generated Excel file
        """
        # Get job
        job = self.db.query(ExtractionJob).filter_by(id=job_id).first()
        if not job:
            raise ValueError(f"Job {job_id} not found")

        # Query results
        query = self.db.query(ExtractionResult).filter_by(job_id=job_id)

        # Apply quality filter
        if min_quality is not None:
            query = query.filter(ExtractionResult.quality_score >= min_quality)

        # Order by quality (best first)
        results = query.order_by(ExtractionResult.quality_score.desc()).all()

        if not results:
            raise ValueError(f"No results found for job {job_id}")

        # Create workbook
        wb = Workbook()

        # Sheet 1: Main data
        ws_data = wb.active
        ws_data.title = "Extraction Results"
        self._populate_data_sheet(ws_data, results, include_metadata, job)

        # Sheet 2: Summary statistics
        ws_summary = wb.create_sheet("Summary")
        self._populate_summary_sheet(ws_summary, job, results)

        # Sheet 3: Data quality report
        if include_metadata:
            ws_quality = wb.create_sheet("Quality Report")
            self._populate_quality_sheet(ws_quality, results)

        # Save file
        filename = self._generate_filename(job)
        filepath = os.path.join(settings.EXPORTS_DIR, filename)

        # Ensure exports directory exists
        os.makedirs(settings.EXPORTS_DIR, exist_ok=True)

        wb.save(filepath)

        logger.info(f"Exported {len(results)} records to {filepath}")

        # Update job with file path
        job.result_file_path = filepath
        self.db.commit()

        return filepath

    def _populate_data_sheet(
        self,
        ws,
        results: List[ExtractionResult],
        include_metadata: bool,
        job: ExtractionJob
    ) -> None:
        """Populate main data sheet with extraction results"""

        # Define columns based on sector
        base_columns = [
            "Name",
            "Email",
            "Phone",
            "Company",
            "Title",
            "Location",
            "LinkedIn URL"
        ]

        metadata_columns = [
            "Quality Score",
            "Confidence Score",
            "Completeness Score",
            "Source URL",
            "Extraction Layer"
        ] if include_metadata else []

        all_columns = base_columns + metadata_columns

        # Write header row
        for col_idx, column_name in enumerate(all_columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)

            # Header styling
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

        # Write data rows
        for row_idx, result in enumerate(results, start=2):
            data = result.data

            # Base data
            values = [
                data.get("name", ""),
                data.get("email", ""),
                data.get("phone", ""),
                data.get("company", ""),
                data.get("title", ""),
                data.get("location", ""),
                data.get("linkedin_url", "")
            ]

            # Metadata
            if include_metadata:
                values.extend([
                    result.quality_score,
                    result.confidence_score,
                    result.completeness_score,
                    result.source_url or "",
                    result.extraction_layer
                ])

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

                # Alignment
                if col_idx > len(base_columns):
                    # Metadata columns - center aligned
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

                # Border
                cell.border = Border(
                    left=Side(style="thin", color="D3D3D3"),
                    right=Side(style="thin", color="D3D3D3"),
                    top=Side(style="thin", color="D3D3D3"),
                    bottom=Side(style="thin", color="D3D3D3")
                )

            # Color-code quality score
            if include_metadata:
                quality_cell = ws.cell(row=row_idx, column=len(base_columns) + 1)
                quality = result.quality_score

                if quality >= 0.8:
                    quality_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                    quality_cell.font = Font(color="006100")
                elif quality >= 0.6:
                    quality_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
                    quality_cell.font = Font(color="9C5700")
                else:
                    quality_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
                    quality_cell.font = Font(color="9C0006")

        # Auto-adjust column widths
        for col_idx, column_name in enumerate(all_columns, start=1):
            column_letter = get_column_letter(col_idx)

            # Set reasonable widths
            if "URL" in column_name or "Source" in column_name:
                ws.column_dimensions[column_letter].width = 50
            elif "Score" in column_name:
                ws.column_dimensions[column_letter].width = 15
            else:
                ws.column_dimensions[column_letter].width = 20

        # Freeze header row
        ws.freeze_panes = "A2"

    def _populate_summary_sheet(
        self,
        ws,
        job: ExtractionJob,
        results: List[ExtractionResult]
    ) -> None:
        """Populate summary statistics sheet"""

        # Title
        ws["A1"] = "Extraction Job Summary"
        ws["A1"].font = Font(bold=True, size=14)

        # Job info
        info = [
            ["Job ID", job.id],
            ["Sector", job.sector.value.capitalize()],
            ["Status", job.status.value.capitalize()],
            ["Started At", job.started_at.strftime("%Y-%m-%d %H:%M:%S") if job.started_at else "N/A"],
            ["Completed At", job.completed_at.strftime("%Y-%m-%d %H:%M:%S") if job.completed_at else "N/A"],
            ["Duration", f"{job.duration_seconds} seconds"],
            ["Total Sources", job.total_sources],
            ["Processed Sources", job.processed_sources],
        ]

        row = 3
        for label, value in info:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1

        # Statistics
        row += 2
        ws[f"A{row}"] = "Extraction Statistics"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        total_records = len(results)
        valid_records = len([r for r in results if r.is_validated])
        high_quality = len([r for r in results if r.quality_score >= 0.8])
        medium_quality = len([r for r in results if 0.6 <= r.quality_score < 0.8])
        low_quality = len([r for r in results if r.quality_score < 0.6])

        avg_quality = sum(r.quality_score for r in results) / total_records if results else 0
        avg_completeness = sum(r.completeness_score for r in results) / total_records if results else 0

        stats = [
            ["Total Records Extracted", total_records],
            ["Valid Records", valid_records],
            ["High Quality (≥80%)", high_quality],
            ["Medium Quality (60-79%)", medium_quality],
            ["Low Quality (<60%)", low_quality],
            ["Average Quality Score", f"{avg_quality:.2%}"],
            ["Average Completeness", f"{avg_completeness:.2%}"],
            ["Duplicate Count", job.duplicate_count],
            ["Error Count", job.error_count]
        ]

        for label, value in stats:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1

        # Auto-adjust columns
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def _populate_quality_sheet(
        self,
        ws,
        results: List[ExtractionResult]
    ) -> None:
        """Populate quality report sheet"""

        # Title
        ws["A1"] = "Data Quality Report"
        ws["A1"].font = Font(bold=True, size=14)

        # Quality distribution
        row = 3
        ws[f"A{row}"] = "Quality Distribution"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 2

        # Header
        ws[f"A{row}"] = "Quality Range"
        ws[f"B{row}"] = "Count"
        ws[f"C{row}"] = "Percentage"

        for col in ["A", "B", "C"]:
            ws[f"{col}{row}"].font = Font(bold=True)
            ws[f"{col}{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws[f"{col}{row}"].font = Font(bold=True, color="FFFFFF")

        row += 1

        # Calculate distribution
        total = len(results)
        ranges = [
            ("90-100%", 0.9, 1.0),
            ("80-89%", 0.8, 0.9),
            ("70-79%", 0.7, 0.8),
            ("60-69%", 0.6, 0.7),
            ("50-59%", 0.5, 0.6),
            ("<50%", 0.0, 0.5)
        ]

        for range_label, min_score, max_score in ranges:
            count = len([r for r in results if min_score <= r.quality_score < max_score])
            percentage = (count / total * 100) if total > 0 else 0

            ws[f"A{row}"] = range_label
            ws[f"B{row}"] = count
            ws[f"C{row}"] = f"{percentage:.1f}%"
            row += 1

        # Field completeness analysis
        row += 2
        ws[f"A{row}"] = "Field Completeness"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 2

        ws[f"A{row}"] = "Field"
        ws[f"B{row}"] = "Populated Count"
        ws[f"C{row}"] = "Percentage"

        for col in ["A", "B", "C"]:
            ws[f"{col}{row}"].font = Font(bold=True)

        row += 1

        fields = ["name", "email", "phone", "company", "title", "location", "linkedin_url"]

        for field in fields:
            count = len([r for r in results if r.data.get(field)])
            percentage = (count / total * 100) if total > 0 else 0

            ws[f"A{row}"] = field.replace("_", " ").title()
            ws[f"B{row}"] = count
            ws[f"C{row}"] = f"{percentage:.1f}%"
            row += 1

        # Auto-adjust columns
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15

    def _generate_filename(self, job: ExtractionJob) -> str:
        """Generate unique filename for export"""
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        sector = job.sector.value
        return f"extraction_{sector}_{job.id}_{timestamp}.xlsx"
